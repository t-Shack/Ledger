"""
Statement parser.

Turns an uploaded bank statement (CSV, XLSX, or PDF) into a flat list of
ParsedTransaction records, regardless of which bank issued it. Column
headers vary by bank ("Narration" vs "Description", separate "Debit"/
"Credit" columns vs a single signed "Amount"), so parsing goes through a
ColumnMapper that matches headers by alias rather than assuming a fixed
layout.

Malformed rows are skipped and reported rather than failing the whole
statement -- one bad row (e.g. a subtotal line in a PDF export) shouldn't
block the rest of the import.

Dependencies: pdfplumber, openpyxl, python-dateutil
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Literal

import pdfplumber
from dateutil import parser as date_parser
from openpyxl import load_workbook

Direction = Literal["debit", "credit"]


@dataclass(frozen=True)
class ParsedTransaction:
    txn_date: date
    raw_description: str
    amount: Decimal
    direction: Direction
    balance_after: Decimal | None = None


@dataclass(frozen=True)
class SkippedRow:
    row_number: int
    raw: list[str]
    reason: str


@dataclass
class ParseResult:
    transactions: list[ParsedTransaction] = field(default_factory=list)
    skipped: list[SkippedRow] = field(default_factory=list)


class StatementParseError(Exception):
    """Raised when a file can't be parsed at all: wrong format, empty, or no
    recognizable header row. Distinct from a SkippedRow, which is a single
    bad line in an otherwise-parseable file.
    """


# ---------------------------------------------------------------------------
# Column mapping: header name -> canonical field, matched by alias substring
# rather than exact string, since every bank names its columns differently.
# ---------------------------------------------------------------------------

_FIELD_ALIASES: dict[str, list[str]] = {
    "date": ["date", "value date", "transaction date", "txn date"],
    "description": ["description", "narration", "detail", "remarks", "particulars"],
    "debit": ["debit", "withdrawal", "money out"],
    "credit": ["credit", "deposit", "money in"],
    "amount": ["amount"],
    "type": ["type", "transaction type", "dr/cr"],
    "balance": ["balance", "running balance", "closing balance"],
}

# Real statements mark an empty debit/credit cell with a placeholder rather
# than leaving it blank -- Opay and Access Bank both use "-"/"--" in the
# statements this was tested against. Treated as empty, same as "".
_EMPTY_PLACEHOLDERS = {"-", "--", "—", "–", "n/a", "na", "nil", "null"}


class ColumnMapper:
    """Maps a statement's actual header row to canonical fields."""

    def __init__(self, headers: list[str]):
        normalized = [h.strip().lower() for h in headers]
        self.index: dict[str, int] = {}
        for field_name, aliases in _FIELD_ALIASES.items():
            for i, h in enumerate(normalized):
                if any(alias in h for alias in aliases):
                    self.index[field_name] = i
                    break

        has_debit_credit = "debit" in self.index and "credit" in self.index
        has_amount = "amount" in self.index
        if "date" not in self.index:
            raise StatementParseError(f"Could not find a date column in headers: {headers}")
        if "description" not in self.index and "type" not in self.index:
            raise StatementParseError(
                f"Could not find a description or transaction-type column in headers: {headers}"
            )
        if not has_debit_credit and not has_amount:
            raise StatementParseError(f"Could not find debit/credit or amount columns in headers: {headers}")
        self.uses_debit_credit = has_debit_credit

    def get(self, row: list[str], field_name: str) -> str | None:
        i = self.index.get(field_name)
        if i is None or i >= len(row):
            return None
        value = row[i].strip()
        if not value or value.lower() in _EMPTY_PLACEHOLDERS:
            return None
        return value


_CURRENCY_NOISE = re.compile(r"[₦$€£,\s]")

# Generous upper bound for a single personal-banking transaction. Exists to
# catch a misaligned row where a long numeric Transaction ID (often 15+
# digits) lands in an amount column -- Decimal() parses it without error,
# so without this check it would silently corrupt totals rather than fail
# to parse.
_MAX_PLAUSIBLE_AMOUNT = Decimal("1000000000000")  # 1 trillion


def _parse_amount(raw: str) -> Decimal:
    cleaned = _CURRENCY_NOISE.sub("", raw)
    negative = cleaned.startswith("(") and cleaned.endswith(")")
    cleaned = cleaned.strip("()")
    value = Decimal(cleaned)
    result = -value if negative else value
    if abs(result) > _MAX_PLAUSIBLE_AMOUNT:
        raise ValueError(f"implausible amount, likely a misread reference number: {raw!r}")
    return result


_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _parse_txn_date(raw: str) -> date:
    """Parses a transaction date, preferring the unambiguous ISO YYYY-MM-DD
    reading when present. dayfirst=True is applied only to genuinely
    ambiguous formats (Nigerian statements commonly use DD/MM/YYYY) --
    applying it to an already-unambiguous ISO date silently swaps day and
    month whenever the day is <= 12 (2026-07-01 becomes January 7th).
    """
    stripped = raw.strip()
    if _ISO_DATE.match(stripped):
        return date_parser.parse(stripped, dayfirst=False).date()
    return date_parser.parse(stripped, dayfirst=True).date()


# Word-boundary patterns, checked in order, for inferring direction from a
# transaction-type word when there's no separate debit/credit column.
# Boundaries matter: a bare "in" would otherwise match inside "Interest".
_DEBIT_TYPE_PATTERNS = [
    r"\bdebit\b", r"\bdr\b", r"\bwithdrawal\b", r"\btransfer out\b",
    r"\bpurchase\b", r"\bpayment\b", r"\bsent\b", r"\bout\b",
]
_CREDIT_TYPE_PATTERNS = [
    r"\bcredit\b", r"\bcr\b", r"\bdeposit\b", r"\btransfer in\b",
    r"\breceived\b", r"\brefund\b", r"\bbonus\b", r"\bin\b",
]


def _infer_direction_from_type(raw_type: str) -> Direction | None:
    """Guesses direction from a transaction-type word/phrase, covering
    conventions beyond literal "debit"/"credit" -- e.g. Withdrawal/Deposit,
    Sent/Received, or bare IN/OUT (the exact wording PalmPay's own app uses).
    Returns None if nothing recognizable is found, so the caller can fall
    back to the amount's sign.
    """
    text = raw_type.lower()
    if any(re.search(p, text) for p in _DEBIT_TYPE_PATTERNS):
        return "debit"
    if any(re.search(p, text) for p in _CREDIT_TYPE_PATTERNS):
        return "credit"
    return None


def _row_to_transaction(row: list[str], mapper: ColumnMapper) -> ParsedTransaction:
    raw_date = mapper.get(row, "date")
    raw_type = mapper.get(row, "type")
    # Some real exports (e.g. Palmpay's) have no narration column at all --
    # the transaction type ("Withdrawal", "Bill Payment") is the closest
    # thing to a description, so it's used when nothing better exists.
    raw_description = mapper.get(row, "description") or raw_type
    if not raw_date or not raw_description:
        raise ValueError("missing date or description")

    txn_date = _parse_txn_date(raw_date)

    if mapper.uses_debit_credit:
        debit = mapper.get(row, "debit")
        credit = mapper.get(row, "credit")
        if debit:
            amount, direction = abs(_parse_amount(debit)), "debit"
        elif credit:
            amount, direction = abs(_parse_amount(credit)), "credit"
        else:
            raise ValueError("no debit or credit value present")
    else:
        raw_amount = mapper.get(row, "amount")
        if raw_amount is None:
            raise ValueError("missing amount")
        amount = _parse_amount(raw_amount)
        inferred = _infer_direction_from_type(raw_type or "")
        direction = inferred if inferred is not None else ("debit" if amount < 0 else "credit")
        amount = abs(amount)

    raw_balance = mapper.get(row, "balance")
    balance_after = _parse_amount(raw_balance) if raw_balance else None

    return ParsedTransaction(
        txn_date=txn_date,
        raw_description=raw_description,
        amount=amount,
        direction=direction,
        balance_after=balance_after,
    )


def _rows_to_result(mapper: ColumnMapper, rows: list[list[str]], row_offset: int = 2) -> ParseResult:
    result = ParseResult()
    for i, row in enumerate(rows, start=row_offset):
        if not any(cell.strip() for cell in row):
            continue  # blank line
        try:
            result.transactions.append(_row_to_transaction(row, mapper))
        except (ValueError, InvalidOperation) as e:
            result.skipped.append(SkippedRow(row_number=i, raw=row, reason=str(e)))
    return result


# ---------------------------------------------------------------------------
# Format-specific readers. Each produces (header, rows) or a sequence of
# tables, and hands off to _rows_to_result, so the row-level logic above is
# shared across all formats.
# ---------------------------------------------------------------------------

def parse_csv(content: bytes) -> ParseResult:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if any(cell.strip() for cell in r)]
    if not rows:
        raise StatementParseError("CSV file is empty")
    mapper = ColumnMapper(rows[0])
    return _rows_to_result(mapper, rows[1:])


def parse_xlsx(content: bytes) -> ParseResult:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [[str(cell) if cell is not None else "" for cell in row] for row in sheet.iter_rows(values_only=True)]
    rows = [r for r in rows if any(cell.strip() for cell in r)]
    if not rows:
        raise StatementParseError("Spreadsheet is empty")
    mapper = ColumnMapper(rows[0])
    return _rows_to_result(mapper, rows[1:])


def _trim_empty_edge_columns(table: list[list[str]]) -> list[list[str]]:
    """Removes leading/trailing columns that are empty across every row.
    pdfplumber's line-based detection sometimes adds phantom empty columns
    at a table's edges, and does so inconsistently page-to-page -- the same
    logical table can come back as 8 columns on one page and 12 on the
    next, purely from this padding. Trimming normalizes both to the same
    width so continuation pages are recognized as continuations.
    """
    if not table:
        return table
    width = max(len(row) for row in table)
    padded = [list(row) + [""] * (width - len(row)) for row in table]

    def col_is_empty(idx: int) -> bool:
        return all(not (row[idx] or "").strip() for row in padded)

    left = 0
    while left < width and col_is_empty(left):
        left += 1
    right = width
    while right > left and col_is_empty(right - 1):
        right -= 1

    return [row[left:right] for row in padded]


def parse_pdf(content: bytes) -> ParseResult:
    """Extracts every table pdfplumber finds across all pages, but doesn't
    trust the first one blindly -- real statements have non-transaction
    "tables" too (account-info boxes, summary blocks) that pdfplumber's
    line-based detection can pick up before or alongside the real one.

    Each table is trimmed of empty edge columns, then searched row-by-row
    for a valid transaction header rather than assuming row 0 is it --
    some exports (e.g. PalmPay's) merge an account-info box into the same
    detected table as the header, ahead of it. A table with no header of
    its own is tried against the most recently established header instead
    of requiring an exact column-count match -- a page containing only
    debits (or only credits) can lose that column entirely from its
    extraction, not just leave it empty, so width isn't a reliable gate.
    Rows that don't actually fit still fail per-row validation and land in
    skipped, same as any other malformed row.
    """
    combined = ParseResult()
    current_mapper: ColumnMapper | None = None
    found_any_table = False

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            for raw_table in page.extract_tables():
                if not raw_table:
                    continue
                table = _trim_empty_edge_columns([[c or "" for c in row] for row in raw_table])
                if not table:
                    continue

                mapper = None
                body = table
                for idx, row in enumerate(table):
                    try:
                        mapper = ColumnMapper(row)
                        body = table[idx + 1:]
                        break
                    except StatementParseError:
                        continue

                if mapper is None:
                    if current_mapper is None:
                        continue  # no established header yet -- nothing to fall back to
                    mapper = current_mapper
                    body = table  # no header row of its own to drop

                current_mapper = mapper
                found_any_table = True

                page_result = _rows_to_result(mapper, body)
                combined.transactions.extend(page_result.transactions)
                combined.skipped.extend(page_result.skipped)

    if not found_any_table:
        raise StatementParseError("No recognizable transaction table found in PDF")
    return combined


def parse_statement(filename: str, content: bytes) -> ParseResult:
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext == "csv":
        return parse_csv(content)
    if ext in ("xlsx", "xls"):
        return parse_xlsx(content)
    if ext == "pdf":
        return parse_pdf(content)
    raise StatementParseError(f"Unsupported file type: .{ext}")
